"""Excel helpers: import row reading and export writing (manual 11).

Exports use Chinese headers, freeze the first row, enable auto-filter and set
text format for id-like columns so values such as 工号/账号/MAC/IP are not
mangled by spreadsheet auto-conversion (manual 11.5).
"""
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Columns that must be written/read as text to preserve leading zeros etc.
TEXT_COLUMNS = {
    "identity_no", "account_name", "ip_address", "mac_address",
    "contact_phone", "phone", "switch_port",
}


def cell_to_text(value: Any) -> str:
    """Best-effort conversion of a raw cell value to a clean string."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def read_rows(path: str) -> tuple[list[str], list[dict]]:
    """Read the active sheet into a list of header-keyed row dicts."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not raw:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in raw[0]]
    data: list[dict] = []
    for r in raw[1:]:
        if r is None or all(c is None for c in r):
            continue
        row: dict = {}
        for i, h in enumerate(headers):
            row[h] = r[i] if i < len(r) else None
        data.append(row)
    return headers, data


def write_export(
    headers: dict[str, str],
    rows: list[dict],
    path: str,
    text_columns: set[str] | None = None,
) -> None:
    """Write an export workbook with Chinese headers, freeze + autofilter."""
    text_columns = text_columns or TEXT_COLUMNS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "export"
    cols = list(headers.keys())
    for c, key in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=headers[key])
        cell.font = Font(bold=True)
    for r, item in enumerate(rows, start=2):
        for c, key in enumerate(cols, start=1):
            val = item.get(key)
            if val is None:
                val = ""
            if key in text_columns:
                val = cell_to_text(val)
            cell = ws.cell(row=r, column=c, value=val)
            if key in text_columns:
                cell.number_format = "@"
    ws.freeze_panes = "A2"
    if rows:
        last_col = get_column_letter(len(cols))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    for c, key in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(
            12, min(40, len(str(headers[key])) + 6)
        )
    wb.save(path)
